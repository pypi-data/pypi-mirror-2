from pyf.componentized.components.multiwriter import MultipleFileWriter

from pyf.componentized.configuration.keys import RepeatedKey, SimpleKey,\
    CompoundKey
from pyf.componentized.configuration.fields import InputField
import operator

from openpyxl import workbook
from openpyxl.cell import get_column_letter
from openpyxl.writer import excel

class XLSXWriter(MultipleFileWriter):
    name = "xlsxwriter"
    
    configuration = [SimpleKey('target_filename', label="Target filename", default="filename.xlsx"),
                     RepeatedKey('columns', 'column',
                                 content=CompoundKey('column',
                                                     text_value='title',
                                                     attributes={'attribute': 'attribute',
                                                                 'renderer': 'renderer'},
                                                     fields=[InputField('title', label="Title",
                                                                        help_text="Column title in header"),
                                                             InputField('attribute', label="Attribute",
                                                                        help_text="Source object attribute"),
                                                             InputField('renderer', label="Renderer",
                                                                        help_text="Renderer eval (optionnal)")]))]
    
    def __init__(self, config_node, component_id):
        """Initialize a new CSVWriter
        @param config: SafeConfigParser
        @type config: SafeConfigParser instance

        @param process_name: The process name to use
        @type process_name: String
        """
        self.config_node = config_node
        self.id = component_id

        self.columns = self.get_config_key('columns')

        self.column_definitions = list()
        self.renderers = dict()
        
        self.__retrieve_columns_definition()

    def __retrieve_columns_definition(self):
        default_renderer = lambda value: value
        
        for column_def in self.columns:
            if column_def.get('attribute'):
                attr = column_def.get('attribute')
            else:
                attr = column_def.get('title')
                
            self.column_definitions.append(dict(
                attribute=attr,
                title=column_def.get('title'),
                renderer=column_def.get('renderer') and\
                         (lambda value, rd=column_def.get('renderer'): eval(rd))
                         or default_renderer))

    def get_column_names(self):
        return map(operator.itemgetter('title'), self.column_definitions)
    
    def write(self, values, key, output_filename, target_filename):
        #target_file = open(output_filename, 'wb+')

#        csvengine = DictWriter(target_file,
#                self.get_column_names(), dialect='%s_dialect' % self.id)
        wb = workbook.Workbook()
        worksheet = wb.get_active_sheet()
        
        def write_line(worksheete, line_num, data):
            for column, item in enumerate(data):
                coord = '%s%s' % (get_column_letter(column+1), line_num)
                if item == 0:
                    item = '0'
                    
                worksheete.cell(coord).value = item
        
        write_line(worksheet, 1, self.get_column_names())

        for line_num, data in enumerate(values):
            dat_cols = [getattr(data, col_def['attribute']) for col_def in self.column_definitions]
            dat = [self.column_definitions[colidx]['renderer'](it) for colidx, it in enumerate(dat_cols)]
            write_line(worksheet, line_num+2, dat)
            yield True

        w = excel.ExcelWriter(wb)
        try:
            w.save(output_filename)
        except Exception, e:
            raise
        
        del w
        del worksheet
        del wb
        
        yield True
